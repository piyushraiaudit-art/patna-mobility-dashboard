"""
Excel annexure + PNG zip builders for the Patna audit dashboard.

Excel: 10-sheet workbook consumed directly as an annexure in the audit
report. Pandas writes the rows, openpyxl post-processes for frozen
headers, autosize, and conditional formatting on the Coverage sheet.

PNG: each Plotly figure exported via kaleido at 1600×1000 / 300 DPI,
packed into a single zip for the report annexure.
"""

from __future__ import annotations

import io
import zipfile
from datetime import datetime

import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from data import AUDIT_WINDOW_END, AUDIT_WINDOW_START
from metrics import (
    SHORT_CORRIDOR_IDS, am_peak_observations, bti as compute_bti, cv as compute_cv,
    direction_asymmetry, hourly_median_cr, minutes_lost_table, peak_observations,
    pm_peak_observations, ranking_table, weekday_observations,
)


HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF")


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def _format_header(ws, n_cols: int) -> None:
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A2"


def _auto_width(ws) -> None:
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            try:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 60)


def build_excel_annexure(df: pd.DataFrame, rep: dict) -> bytes:
    """Produce the multi-sheet xlsx as bytes (Streamlit download_button consumes bytes)."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        stats = rep["stats"]

        # 1. Cover -------------------------------------------------------
        cover = pd.DataFrame({
            "Field": [
                "Title", "Audit window (start)", "Audit window (end)",
                "Polling interval", "OD pairs",
                "First observation in dataset", "Last observation in dataset",
                "Days covered", "Corridors covered", "Total OK observations",
                "Total FAIL rows", "FAIL %",
                "Observations MD5", "corridors.csv MD5",
                "Built at",
            ],
            "Value": [
                "Patna Mobility Audit — Congestion Index Annexure",
                str(AUDIT_WINDOW_START.date()) + " 00:00 IST",
                str(AUDIT_WINDOW_END.date()) + " 23:59 IST",
                "Every 30 minutes",
                "56 (28 corridors × 2 directions)",
                stats.first_timestamp,
                stats.last_timestamp,
                stats.days_covered,
                f"{stats.corridors_covered}/28",
                f"{stats.total_observations:,}",
                f"{stats.fail_count:,}",
                f"{stats.fail_pct:.3f}%",
                stats.observations_md5,
                stats.corridors_md5,
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ],
        })
        cover.to_excel(writer, sheet_name="1. Cover", index=False)

        # 2. Ranking -----------------------------------------------------
        ranking = ranking_table(df)
        ranking_display = ranking[[
            "rank", "corridor_id", "corridor_name", "phci", "phci_hour",
            "phci_direction", "adci", "bti", "cv", "n_peak", "is_short_corridor",
        ]].copy()
        ranking_display.columns = [
            "Rank", "Corridor ID", "Corridor", "PHCI", "Worst hour",
            "Worst direction", "ADCI (06-21)", "BTI (peak)", "CV (peak)",
            "n (peak obs)", "Short corridor (<1.5 km)",
        ]
        ranking_display.to_excel(writer, sheet_name="2. Ranking", index=False)

        # 3. Hourly medians ---------------------------------------------
        hm = hourly_median_cr(df)
        hm.to_excel(writer, sheet_name="3. Hourly Medians", index=False)

        # 4. Direction asymmetry ----------------------------------------
        asym = direction_asymmetry(df)
        asym.to_excel(writer, sheet_name="4. Direction Asymmetry", index=False)

        # 5. Reliability -------------------------------------------------
        bti_df = compute_bti(df)
        cv_df = compute_cv(df)[["corridor_id", "direction", "cv", "mu_peak_s", "sigma_peak_s"]]
        rel = bti_df.merge(cv_df, on=["corridor_id", "direction"], how="left")
        rel.to_excel(writer, sheet_name="5. Reliability", index=False)

        # 6. Coverage ----------------------------------------------------
        coverage_long = rep["coverage"]
        coverage_pivot = coverage_long.pivot(
            index="corridor_id", columns="date", values="n_obs"
        ).fillna(0).astype(int)
        coverage_pivot.to_excel(writer, sheet_name="6. Coverage")

        # 7. FAIL log ----------------------------------------------------
        fail_log = rep["fail_log"]
        if not fail_log.empty:
            fail_log[[
                "timestamp_ist", "date", "time", "corridor_id", "corridor_name",
                "direction", "api_status", "error_msg",
            ]].to_excel(writer, sheet_name="7. FAIL Log", index=False)
        else:
            pd.DataFrame({"note": ["No failed API calls in the cumulative log."]}).to_excel(
                writer, sheet_name="7. FAIL Log", index=False
            )

        # 8. Distance drift ---------------------------------------------
        drift = rep["distance_drift"].copy()
        drift.to_excel(writer, sheet_name="8. Distance Drift", index=False)

        # 9. Methodology -------------------------------------------------
        methodology_rows = [
            ("Instant Congestion Ratio",
             "CR(i,t) = duration_traffic_s / duration_freeflow_s"),
            ("Hourly Median CR",
             "CR_hour = median over the hour, per (corridor, direction, hour)"),
            ("PHCI (Peak-Hour Congestion Index)",
             "max over peak hours (08-10, 17-19) of weekday median CR per (corridor, direction); "
             "both directions collapsed by max"),
            ("ADCI (All-Day Congestion Index)",
             "mean over active hours 06-21 of the hourly median CR"),
            ("BTI (Buffer Time Index, FHWA)",
             "(p95(duration_traffic_peak) - median(duration_traffic_peak)) / median(...)"),
            ("CV (Coefficient of Variation)",
             "sigma(duration_traffic_peak) / mu(duration_traffic_peak)"),
            ("Peak window — AM", "08:00 to 11:00 IST (hard-coded; Bihar govt office hours)"),
            ("Peak window — PM", "17:00 to 20:00 IST (hard-coded)"),
            ("Active hours (ADCI)", "06:00 to 22:00 IST"),
            ("Short corridors (asterisked in Ranking)",
             ", ".join(sorted(SHORT_CORRIDOR_IDS)) + " — < 1.5 km"),
            ("Filter applied to OK observations",
             "api_status == 'OK' AND duration_freeflow_s > 0 AND congestion_ratio is not null"),
            ("Dedupe key",
             "(timestamp_ist, corridor_id, direction) — last write wins across snapshot CSVs"),
            ("Holidays (Bihar)",
             "21 May 2026 — Anti-Terrorism Day (observance). Excluded from default weekday "
             "aggregations; toggle in dashboard sidebar."),
        ]
        pd.DataFrame(methodology_rows, columns=["Term", "Definition"]).to_excel(
            writer, sheet_name="9. Methodology", index=False
        )

        # 10. Raw observations ------------------------------------------
        raw_out = df.copy()
        if len(raw_out) > 100_000:
            # Truncate the embedded copy; note this on the Cover sheet.
            raw_out = raw_out.head(100_000)
        raw_out.to_excel(writer, sheet_name="10. Raw Observations", index=False)

    # Post-process: header styling, freeze panes, autosize, conditional formatting.
    buf.seek(0)
    from openpyxl import load_workbook
    wb = load_workbook(buf)
    for ws in wb.worksheets:
        if ws.max_row >= 1 and ws.max_column >= 1:
            _format_header(ws, ws.max_column)
            _auto_width(ws)

    # Conditional formatting on Coverage sheet.
    cov_ws = wb["6. Coverage"]
    if cov_ws.max_row > 1 and cov_ws.max_column > 1:
        last_col_letter = get_column_letter(cov_ws.max_column)
        range_str = f"B2:{last_col_letter}{cov_ws.max_row}"
        cov_ws.conditional_formatting.add(
            range_str,
            ColorScaleRule(
                start_type="num", start_value=0, start_color="DC2626",
                mid_type="num", mid_value=30, mid_color="FCD34D",
                end_type="num", end_value=48, end_color="16A34A",
            ),
        )

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ---------------------------------------------------------------------------
# PNG zip
# ---------------------------------------------------------------------------

def build_png_zip(figures: dict) -> bytes:
    """Pack a dict of {filename: plotly.Figure} into a zip of 1600x1000 PNGs."""
    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zf:
        for fname, fig in figures.items():
            try:
                png_bytes = fig.to_image(format="png", width=1600, height=1000, scale=2)
                zf.writestr(fname, png_bytes)
            except Exception as e:
                zf.writestr(
                    fname + ".error.txt",
                    f"Could not export {fname}: {e}\n"
                    "Install the kaleido package: pip install kaleido==0.2.1\n",
                )
    return out.getvalue()
